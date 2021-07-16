

import torch
from utils.common import *
from utils.log import get_logger
from torchvision import models
from data_selection import ScoreCalciumSelection
import torch.nn as nn
from utils import statistics, dataset, train, test
import time
import matplotlib.pyplot as plt
from datetime import datetime
from utils.classifiers import run_classifiers

# Login instance
log = get_logger()

######################################
# LOCAL HYPER-PARAMETERS             #
BATCH_SIZE = 8
EPOCHS = 1
LEARNING_RATE = 0.0001
N_CLASSES = 2
MONOFOLD = True
######################################

def get_model():
    '''

    :return:
    '''
    base_model = 'tmp/base_model.pt'
    net = models.vgg16(pretrained=True)

    # Freeze trained weights
    for param in net.features.parameters():
        param.requires_grad = False

    # Newly created modules have require_grad=True by default
    num_features = net.classifier[6].in_features
    features = list(net.classifier.children())[:-1]  # Remove last layer
    # torch.manual_seed(3)
    linear = nn.Linear(num_features, N_CLASSES)
    torch.nn.init.xavier_uniform(linear.weight)
    linear.bias.data.fill_(0.01)

    print(linear.weight.sum())
    features.extend([linear])  # Add our layer with 2 outputs
    net.classifier = nn.Sequential(*features)  # Replace the model classifier

    # if os.path.exists(base_model):
    #     net.load_state_dict(torch.load(base_model))
    #     log.info(f'loading {base_model}')
    # else:
    #     torch.save(net.state_dict(), 'tmp/base_model.pt')
    #     log.info(f'save {base_model}')

    net.to(device=device)

    return net


def get_excel_file(infile, outfile, data, guide):
    import openpyxl
    wb = openpyxl.reader.excel.load_workbook(infile)


    added = {}
    for sheet in wb.worksheets:
        sheet['E1'] = 'id_l'
        sheet['F1'] = 'P_l'
        sheet['G1'] = 'SF1_l'
        sheet['H1'] = 'SF2_l'
        sheet['I1'] = 'gr_l'
        sheet['J1'] = 'f_l'
        sheet['K1'] = 'cls_l'

        sheet['L1'] = 'id_r'
        sheet['M1'] = 'P_r'
        sheet['N1'] = 'SF1_r'
        sheet['O1'] = 'SF2_r'
        sheet['P1'] = 'gr_r'
        sheet['Q1'] = 'f_r'
        sheet['R1'] = 'cls_r'

        added[sheet.title] = []
    for cls, samples in guide.items():
        for sample in samples:
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(sheet.min_row, sheet.max_row):
                    if sample[0:-3] == row[0].value:
                        for elem in data:
                            if elem[0][0:-3] == row[0].value and elem[6] == cls:
                                if elem[0].endswith('OIc'):
                                    added[sheet.title].append(row[0].row)
                                    sheet[f'E{row[0].row}'] = elem[0]
                                    sheet[f'F{row[0].row}'] = elem[1]
                                    sheet[f'G{row[0].row}'] = elem[2]
                                    sheet[f'H{row[0].row}'] = elem[3]
                                    sheet[f'I{row[0].row}'] = elem[4]
                                    sheet[f'J{row[0].row}'] = elem[5]
                                    sheet[f'K{row[0].row}'] = elem[6]
                                elif elem[0].endswith('ODc'):
                                    added[sheet.title].append(row[0].row)
                                    sheet[f'L{row[0].row}'] = elem[0]
                                    sheet[f'M{row[0].row}'] = elem[1]
                                    sheet[f'N{row[0].row}'] = elem[2]
                                    sheet[f'O{row[0].row}'] = elem[3]
                                    sheet[f'P{row[0].row}'] = elem[4]
                                    sheet[f'Q{row[0].row}'] = elem[5]
                                    sheet[f'R{row[0].row}'] = elem[6]
                                else:
                                    print('not possible')
                added[sheet.title] = list(dict.fromkeys(added[sheet.title]))

    todelete = {}
    for sheet in wb.worksheets:
        todelete[sheet.title] = []
        for row in sheet.iter_rows(sheet.min_row, sheet.max_row):
            if row[0].row in added[sheet.title]:
                continue
            else:
                todelete[sheet.title].append(row[0].row)

    for sheet, rows in todelete.items():
        for r in reversed(rows):
            wb[sheet].delete_rows(r)

    wb.save(outfile)


if __name__ == '__main__':
    """
    Set of experiments that combine IA and clinical data
    Sequence:
        0.- Init models to be common to each experiment
        1.- VGG16 pretrained imagenet
        2.- for fold_id in folds:
            2.1. Train net on train set
            2.2. Test net on train+test set
                2.2.1. Get SF1+SF2 column data (continuous prediction)
                2.2.2. Get model prediction column data (discrete prediction)
                2.2.3. Generate excel combining clinical data and predicted data
            2.3. Train classifier with train+test data (all excel rows)
            2.4. Test classifier with test data (same set used in CNN test step) 
    """
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    log.info(f'Using device: {device}')

    rt = ScoreCalciumSelection('data/folds/', 'data/sc_run', 'data/originals/sc/')

    folds_acc = []
    for i in range(5):
        t0 = time.time()
        train_data, test_data = rt.generate_run_set(i + 1)

        # # Import model
        # model = get_model()
        #
        # log.info(f'Generate test with fold {i + 1}')
        # print(f'Train fold {i + 1}')
        # log.info("Train model")
        #
        # # Load and transform datasets
        # train_data_loader = dataset.load_and_transform_data(os.path.join(SCORE_CALCIUM_DATA, TRAIN), log=log,
        #                                                     batch_size=BATCH_SIZE,
        #                                                     data_augmentation=False)
        #
        # # train model
        # model, test_acc, lrs, epoch_list, losses = train.train_model(model=model,
        #                                                              device=device,
        #                                                              train_loader=train_data_loader,
        #                                                              log=log,
        #                                                              epochs=EPOCHS,
        #                                                              batch_size=BATCH_SIZE,
        #                                                              lr=LEARNING_RATE,
        #                                                              test_loader=None)
        #
        # # # test model
        # # test_data_loader = dataset.load_and_transform_data(os.path.join(SCORE_CALCIUM_DATA, TEST), log=log)
        # # log.info("Test model after training")
        # # acc_model, fold_data = test.evaluate_model(model, test_data_loader, device, log=log, fold=i + 1)
        # #
        # # test_results = [elem[1] for elem in fold_data]
        #
        # # test model on TRAIN + TEST
        # test_data_loader = dataset.load_and_transform_data(os.path.join(SCORE_CALCIUM_DATA, TRAIN_TEST), log=log)
        # log.info("Test model after training")
        # acc_model, fold_data = test.evaluate_model(model, test_data_loader, device, log=log, fold=i + 1)

        fold_data= [('p105ODc', 0, -0.27052453, -0.6572126, 1, 1, 'CACSmas400'),
                    ('p105OIc', 0, -0.3965457, -0.6208172, 1, 1, 'CACSmas400'),
                    ('p106ODc', 0, -0.06665759, -0.13635254, 1, 1, 'CACSmas400'),
                    ('p106OIc', 1, -0.9029724, -0.50871515, 1, 1, 'CACSmas400'),
                    ('p113ODc', 0, -0.24596523, -0.6003495, 1, 1, 'CACSmas400'),
                    ('p113OIc', 1, -0.6991856, -0.41291523, 1, 1, 'CACSmas400'),
                    ('p115ODc', 0, 0.38649052, -1.1958174, 1, 1, 'CACSmas400'),
                    ('p115OIc', 0, -0.010677623, -1.3913658, 1, 1, 'CACSmas400'),
                    ('p118ODc', 0, -0.07457928, -0.64052963, 1, 1, 'CACSmas400'),
                    ('p118OIc', 0, -0.077570245, -0.8146767, 1, 1, 'CACSmas400'),
                    ('p120ODc', 0, 0.24493082, -1.3434669, 1, 1, 'CACSmas400'),
                    ('p120OIc', 0, 0.016499056, -1.2350099, 1, 1, 'CACSmas400'),
                    ('p121ODc', 1, -0.40269578, -0.14449736, 1, 1, 'CACSmas400'),
                    ('p121OIc', 0, -0.22018503, -0.40949735, 1, 1, 'CACSmas400'),
                    ('p126ODc', 0, -0.3082636, -0.60847855, 1, 1, 'CACSmas400'),
                    ('p126OIc', 0, -0.37656182, -0.74684435, 1, 1, 'CACSmas400'),
                    ('p134ODc', 1, -0.84181565, -0.2962893, 1, 1, 'CACSmas400'),
                    ('p134OIc', 0, -0.023415493, -0.7236547, 1, 1, 'CACSmas400'),
                    ('p137ODc', 0, -0.24839745, -1.0427496, 1, 1, 'CACSmas400'),
                    ('p137OIc', 0, -0.23971258, -0.9265262, 1, 1, 'CACSmas400'),
                    ('p142ODc', 0, 0.22561227, -1.2660041, 1, 1, 'CACSmas400'),
                    ('p142OIc', 0, -0.8008582, -0.8913828, 1, 1, 'CACSmas400'),
                    ('p146ODc', 0, 0.06828545, -0.21088049, 1, 1, 'CACSmas400'),
                    ('p146OIc', 0, 0.3445695, -0.8478477, 1, 1, 'CACSmas400'),
                    ('p179ODc', 0, -0.5649128, -1.4040571, 1, 1, 'CACSmas400'),
                    ('p179OIc', 0, -0.16918035, -1.2957668, 1, 1, 'CACSmas400'),
                    ('p194ODc', 0, -0.6421502, -0.83334595, 1, 1, 'CACSmas400'),
                    ('p194OIc', 0, 0.28426117, -1.2649772, 1, 1, 'CACSmas400'),
                    ('p195ODc', 0, -0.45672017, -0.6623994, 1, 1, 'CACSmas400'),
                    ('p195OIc', 1, -0.37207425, -0.08065262, 1, 1, 'CACSmas400'),
                    ('p200ODc', 0, -0.5730131, -0.71573836, 1, 1, 'CACSmas400'),
                    ('p200OIc', 0, -0.49938625, -0.6427719, 1, 1, 'CACSmas400'),
                    ('p201ODc', 0, -0.09404896, -0.44086793, 1, 1, 'CACSmas400'),
                    ('p201OIc', 0, 0.3877275, -1.0968132, 1, 1, 'CACSmas400'),
                    ('p202ODc', 0, -0.07071133, -0.8476861, 1, 1, 'CACSmas400'),
                    ('p202OIc', 0, -0.11810951, -1.506859, 1, 1, 'CACSmas400'),
                    ('p207ODc', 0, -0.20438276, -1.1077197, 1, 1, 'CACSmas400'),
                    ('p207OIc', 0, -0.6582682, -1.169134, 1, 1, 'CACSmas400'),
                    ('p208ODc', 0, -0.38211977, -0.87440306, 1, 1, 'CACSmas400'),
                    ('p208OIc', 0, -0.24434938, -0.5048317, 1, 1, 'CACSmas400'),
                    ('p211ODc', 0, 0.40457857, -1.2652853, 1, 1, 'CACSmas400'),
                    ('p211OIc', 0, -0.15985529, -1.163955, 1, 1, 'CACSmas400'),
                    ('p215ODc', 0, 0.5411215, -0.9110482, 1, 1, 'CACSmas400'),
                    ('p215OIc', 0, -0.07020743, -0.81282794, 1, 1, 'CACSmas400'),
                    ('p216ODc', 0, 0.6861228, -1.5329505, 1, 1, 'CACSmas400'),
                    ('p216OIc', 0, -0.05055152, -0.7055716, 1, 1, 'CACSmas400'),
                    ('p245ODc', 0, -0.11633073, -0.6999089, 1, 1, 'CACSmas400'),
                    ('p245OIc', 1, -0.5390058, -0.34425348, 1, 1, 'CACSmas400'),
                    ('p33ODc', 0, 0.09688206, -1.323474, 1, 1, 'CACSmas400'),
                    ('p33OIc', 0, -0.3098734, -0.64797515, 1, 1, 'CACSmas400'),
                    ('p46ODc', 0, 0.25061655, -0.9997464, 1, 1, 'CACSmas400'),
                    ('p46OIc', 0, 0.16733666, -0.8937113, 1, 1, 'CACSmas400'),
                    ('p50ODc', 0, 0.12004216, -0.6521039, 1, 1, 'CACSmas400'),
                    ('p50OIc', 0, -0.037702605, -1.0366327, 1, 1, 'CACSmas400'),
                    ('p5ODc', 0, 0.41035694, -0.9998266, 1, 1, 'CACSmas400'),
                    ('p5OIc', 0, 0.11504014, -1.0602959, 1, 1, 'CACSmas400'),
                    ('p66ODc', 0, 0.0525008, -0.901231, 1, 1, 'CACSmas400'),
                    ('p66OIc', 0, -0.20570706, -0.92716485, 1, 1, 'CACSmas400'),
                    ('p68ODc', 0, 0.04785408, -1.4588099, 1, 1, 'CACSmas400'),
                    ('p68OIc', 0, 0.23080446, -1.2117536, 1, 1, 'CACSmas400'),
                    ('p69ODc', 0, -0.4516955, -1.1818584, 1, 1, 'CACSmas400'),
                    ('p69OIc', 0, -0.4838444, -1.1021681, 1, 1, 'CACSmas400'),
                    ('p83ODc', 0, 0.2058415, -0.85478735, 1, 1, 'CACSmas400'),
                    ('p83OIc', 0, 0.081532374, -1.2008003, 1, 1, 'CACSmas400'),
                    ('p8ODc', 0, -0.34056157, -0.99312145, 1, 1, 'CACSmas400'),
                    ('p8OIc', 0, -0.5108585, -0.76314545, 1, 1, 'CACSmas400'),
                    ('p97ODc', 0, -0.5072799, -0.74472797, 1, 1, 'CACSmas400'),
                    ('p97OIc', 0, -0.8739455, -1.0282693, 1, 1, 'CACSmas400'),
                    ('p101ODc', 0, 0.36887756, -1.0270473, 0, 1, 'CACSmenos400'),
                    ('p101OIc', 0, 0.28529122, -1.539011, 0, 1, 'CACSmenos400'),
                    ('p102ODc', 0, -0.0471112, -0.87886167, 0, 1, 'CACSmenos400'),
                    ('p102OIc', 0, 0.09184195, -1.0537056, 0, 1, 'CACSmenos400'),
                    ('p110ODc', 0, 0.1816123, -1.1082759, 0, 1, 'CACSmenos400'),
                    ('p110OIc', 0, -0.04743032, -0.7612229, 0, 1, 'CACSmenos400'),
                    ('p117ODc', 0, 0.38864985, -1.247276, 0, 1, 'CACSmenos400'),
                    ('p117OIc', 0, 0.17609234, -0.7880015, 0, 1, 'CACSmenos400'),
                    ('p122ODc', 0, 0.110870734, -1.0712426, 0, 1, 'CACSmenos400'),
                    ('p122OIc', 0, -0.510662, -0.8037087, 0, 1, 'CACSmenos400'),
                    ('p124ODc', 0, -0.23974042, -1.2331588, 0, 1, 'CACSmenos400'),
                    ('p124OIc', 0, 0.039248005, -1.0974355, 0, 1, 'CACSmenos400'),
                    ('p135ODc', 0, -0.10223581, -0.8914101, 0, 1, 'CACSmenos400'),
                    ('p135OIc', 0, -0.626135, -1.3017445, 0, 1, 'CACSmenos400'),
                    ('p136ODc', 0, 0.457046, -1.473206, 0, 1, 'CACSmenos400'),
                    ('p136OIc', 0, 0.37708896, -1.0112627, 0, 1, 'CACSmenos400'),
                    ('p138ODc', 0, 0.5767153, -1.2627578, 0, 1, 'CACSmenos400'),
                    ('p138OIc', 0, 0.11632492, -1.1572106, 0, 1, 'CACSmenos400'),
                    ('p145ODc', 0, 0.12053184, -1.18997, 0, 1, 'CACSmenos400'),
                    ('p145OIc', 0, -0.37146437, -0.606659, 0, 1, 'CACSmenos400'),
                    ('p149ODc', 0, -0.19818608, -0.86768645, 0, 1, 'CACSmenos400'),
                    ('p149OIc', 0, 0.19528297, -0.45074758, 0, 1, 'CACSmenos400'),
                    ('p151ODc', 0, 0.28979918, -0.64905274, 0, 1, 'CACSmenos400'),
                    ('p151OIc', 0, 0.10384576, -0.5851302, 0, 1, 'CACSmenos400'),
                    ('p152ODc', 0, -0.062130556, -0.8166025, 0, 1, 'CACSmenos400'),
                    ('p152OIc', 0, 0.12072687, -0.8274016, 0, 1, 'CACSmenos400'),
                    ('p157ODc', 0, -0.1442651, -0.9622677, 0, 1, 'CACSmenos400'),
                    ('p157OIc', 0, 0.11577876, -0.9131229, 0, 1, 'CACSmenos400'),
                    ('p164ODc', 1, -0.3635668, 0.076230325, 0, 1, 'CACSmenos400'),
                    ('p164OIc', 0, 0.21682094, -1.1748953, 0, 1, 'CACSmenos400'),
                    ('p170ODc', 0, -0.18876065, -0.5263998, 0, 1, 'CACSmenos400'),
                    ('p170OIc', 0, -0.3304823, -0.3761245, 0, 1, 'CACSmenos400'),
                    ('p173ODc', 0, -0.011577832, -1.0417176, 0, 1, 'CACSmenos400'),
                    ('p173OIc', 0, -0.10996656, -0.90947825, 0, 1, 'CACSmenos400'),
                    ('p174ODc', 0, 0.12700863, -1.1276088, 0, 1, 'CACSmenos400'),
                    ('p174OIc', 0, -0.10996656, -0.90947825, 0, 1, 'CACSmenos400'),
                    ('p176ODc', 0, -0.43128347, -0.59073675, 0, 1, 'CACSmenos400'),
                    ('p176OIc', 0, -0.008086728, -0.45779657, 0, 1, 'CACSmenos400'),
                    ('p177ODc', 0, -0.30293098, -0.88930166, 0, 1, 'CACSmenos400'),
                    ('p177OIc', 0, 0.33436003, -0.9846619, 0, 1, 'CACSmenos400'),
                    ('p181ODc', 0, -0.13878159, -0.83824897, 0, 1, 'CACSmenos400'),
                    ('p181OIc', 0, -0.3382817, -0.84251, 0, 1, 'CACSmenos400'),
                    ('p184ODc', 0, -0.26245052, -1.333241, 0, 1, 'CACSmenos400'),
                    ('p184OIc', 0, -0.5579972, -1.4707979, 0, 1, 'CACSmenos400'),
                    ('p189ODc', 0, 0.35311964, -1.1490564, 0, 1, 'CACSmenos400'),
                    ('p189OIc', 0, 0.3024071, -1.3345754, 0, 1, 'CACSmenos400'),
                    ('p191ODc', 0, 0.33882803, -1.0470093, 0, 1, 'CACSmenos400'),
                    ('p191OIc', 0, -0.11527087, -1.058124, 0, 1, 'CACSmenos400'),
                    ('p213ODc', 0, -0.59810185, -1.1403738, 0, 1, 'CACSmenos400'),
                    ('p213OIc', 0, 0.22980057, -0.8997503, 0, 1, 'CACSmenos400'),
                    ('p214ODc', 0, -0.5018388, -1.2820692, 0, 1, 'CACSmenos400'),
                    ('p214OIc', 0, -0.104852185, -1.0144997, 0, 1, 'CACSmenos400'),
                    ('p217ODc', 0, -0.0030884314, -1.3382916, 0, 1, 'CACSmenos400'),
                    ('p217OIc', 0, 0.038482442, -1.5479953, 0, 1, 'CACSmenos400'),
                    ('p221ODc', 0, 0.5154456, -1.3344606, 0, 1, 'CACSmenos400'),
                    ('p221OIc', 0, 0.4036947, -1.6048753, 0, 1, 'CACSmenos400'),
                    ('p229ODc', 0, -0.039451137, -1.3428113, 0, 1, 'CACSmenos400'),
                    ('p229OIc', 0, -0.6341423, -0.72576475, 0, 1, 'CACSmenos400'),
                    ('p231ODc', 0, 0.20324732, -0.91363925, 0, 1, 'CACSmenos400'),
                    ('p231OIc', 0, -0.12397452, -0.88314706, 0, 1, 'CACSmenos400'),
                    ('p233ODc', 0, 0.6976335, -1.1849368, 0, 1, 'CACSmenos400'),
                    ('p233OIc', 0, 0.42146438, -1.1269732, 0, 1, 'CACSmenos400'),
                    ('p244ODc', 0, 0.11808874, -1.0012282, 0, 1, 'CACSmenos400'),
                    ('p244OIc', 0, -0.3725677, -1.1612945, 0, 1, 'CACSmenos400'),
                    ('p254ODc', 0, 0.20918097, -0.6978299, 0, 1, 'CACSmenos400'),
                    ('p254OIc', 0, -0.4109373, -1.7193809, 0, 1, 'CACSmenos400'),
                    ('p31ODc', 0, -0.027071105, -1.5253546, 0, 1, 'CACSmenos400'),
                    ('p31OIc', 0, -0.22950374, -1.111063, 0, 1, 'CACSmenos400'),
                    ('p32ODc', 0, 0.39218348, -1.1853856, 0, 1, 'CACSmenos400'),
                    ('p32OIc', 0, 0.1605425, -1.3683683, 0, 1, 'CACSmenos400'),
                    ('p39ODc', 0, -0.070908055, -1.4737165, 0, 1, 'CACSmenos400'),
                    ('p39OIc', 0, 0.11253975, -1.0634018, 0, 1, 'CACSmenos400'),
                    ('p40ODc', 0, 0.00089703314, -0.8246245, 0, 1, 'CACSmenos400'),
                    ('p40OIc', 0, 0.0031458568, -0.73414516, 0, 1, 'CACSmenos400'),
                    ('p49ODc', 0, -0.032527134, -0.79876506, 0, 1, 'CACSmenos400'),
                    ('p49OIc', 0, -0.0798018, -0.8016644, 0, 1, 'CACSmenos400'),
                    ('p64ODc', 0, -0.09604825, -0.9379197, 0, 1, 'CACSmenos400'),
                    ('p64OIc', 0, 0.2825516, -1.2798289, 0, 1, 'CACSmenos400'),
                    ('p74ODc', 0, 0.24356444, -1.3724711, 0, 1, 'CACSmenos400'),
                    ('p74OIc', 0, -0.4588682, -0.9293897, 0, 1, 'CACSmenos400'),
                    ('p80ODc', 0, 0.5287326, -0.75875694, 0, 1, 'CACSmenos400'),
                    ('p80OIc', 1, -0.42119685, 0.015704758, 0, 1, 'CACSmenos400'),
                    ('p86ODc', 0, 0.008000923, -0.6252243, 0, 1, 'CACSmenos400'),
                    ('p86OIc', 0, -0.12393628, -1.1592656, 0, 1, 'CACSmenos400'),
                    ('p90ODc', 0, 0.6451534, -1.3759257, 0, 1, 'CACSmenos400'),
                    ('p90OIc', 0, 0.30052748, -0.97267514, 0, 1, 'CACSmenos400')]

        # Generate excel with train data
        train_filename = os.path.join(BASE_DATA, f'clinical/run_folder/clinical_train_f{i + 1}.xlsx')
        train_file = get_excel_file(infile=BASE_CLINICAL, outfile=train_filename, data=fold_data, guide=train_data)

        # Generate excel with test data
        test_filename = os.path.join(BASE_DATA, f'clinical/run_folder/clinical_test_f{i + 1}.xlsx')
        test_file = get_excel_file(infile=BASE_CLINICAL, outfile=test_filename, data=fold_data, guide=test_data)

        # Train Classifiers


    #     log.info(
    #         f'Accuracy after training {acc_model}. | [{time.time() - t0}]')
    #     folds_acc.append(acc_model)
    #
    #     if MONOFOLD:
    #         break
    #
    #
    # # Confident interval computation
    # mean, stdev, offset, ci = statistics.get_fold_metrics(folds_acc)
    # log.info(f'******************************************')
    # log.info(f'Model performance:')
    # log.info(f'     Folds Acc.: {folds_acc}')
    # log.info(f'     Mean: {mean}')
    # log.info(f'     Stdev: {stdev}')
    # log.info(f'     Offset: {offset}')
    # log.info(f'     CI:(95%) : {ci}')
    # print(f'******************************************')
    # print(f'Model performance:')
    # print(f'     Folds Acc.: {folds_acc}')
    # print(f'     Mean: {mean}')
    # print(f'     Stdev: {stdev}')
    # print(f'     Offset: {offset}')
    # print(f'     CI:(95%) : {ci}')
