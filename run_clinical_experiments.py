import os

import torch
from utils.common import *
from utils.log import get_logger
from torchvision import models
from data_selection import ScoreCalciumSelection
import torch.nn as nn
from utils import statistics, dataset, train, test
import time
import matplotlib.pyplot as plt
from datetime import datetime
from utils.classifiers import run_classifiers
import pickle

######################################
# LOCAL HYPER-PARAMETERS             #
BATCH_SIZE = 8
EPOCHS = 80
MONOFOLD = False
LEARNING_RATE = 0.0001
N_CLASSES = 2
SAVE_LOSS = True
TRAIN_NET = False
######################################

# Login instance
log = get_logger()


def get_model():
    '''

    :return:
    '''
    base_model = 'tmp/base_model.pt'
    net = models.vgg16(pretrained=True)

    # Freeze trained weights
    for param in net.features.parameters():
        param.requires_grad = False

    # Newly created modules have require_grad=True by default
    num_features = net.classifier[6].in_features
    features = list(net.classifier.children())[:-1]  # Remove last layer
    features.extend([nn.Linear(num_features, N_CLASSES)])  # Add our layer with 2 outputs
    net.classifier = nn.Sequential(*features)  # Replace the model classifier

    if os.path.exists(base_model):
        net.load_state_dict(torch.load(base_model))
        log.info(f'loading {base_model}')
    else:
        torch.save(net.state_dict(), 'tmp/base_model.pt')
        log.info(f'save {base_model}')

    net.to(device=device)

    return net


def get_fold_rows(book, sheet_id, fold_id):
    wb = openpyxl.reader.excel.load_workbook(book)
    sheet = wb.worksheets[sheet_id]

    rows = []
    for columns in sheet.iter_cols(15, 15):
        for cell in columns:
            if cell.value == fold_id:
                rows.append(cell.row)

    wb.close()
    return rows

if __name__ == '__main__':
    """
    Set of experiments that combine image data and clincal data:
    Sequence:
        0.- Init models to be common to each experiment
        1.- VGG16 pretrained imagenet
        2.- Ensembles classifiers of clinical data
        3.- Add extra column with deep learning prediction
        4.- Remove column DR from clinical data 
    """
    if TRAIN_NET:
        # Import device
        device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
        log.info(f'Using device: {device}')

        # Generate run test
        rt = ScoreCalciumSelection()
        folds_acc = []
        fold_predictions = []

        for i in range(5):
            t0 = time.time()
            rt.generate_run_set(i + 1)
            # Import model
            model = get_model()

            log.info(f'Generate test with fold {i + 1}')
            print(f'Train fold {i + 1}')

            log.info("Train model")

            # Load and transform datasets
            test_data_loader = dataset.load_and_transform_data(os.path.join(SCORE_CALCIUM_DATA, TEST), log=log)
            train_data_loader = dataset.load_and_transform_data(os.path.join(SCORE_CALCIUM_DATA, TRAIN), log=log,
                                                                batch_size=BATCH_SIZE,
                                                                data_augmentation=False)

            # train model
            model, test_acc, lrs, epoch_list, losses = train.train_model(model=model,
                                                                         device=device,
                                                                         train_loader=train_data_loader,
                                                                         log=log,
                                                                         epochs=EPOCHS,
                                                                         batch_size=BATCH_SIZE,
                                                                         lr=LEARNING_RATE,
                                                                         test_loader=None)

            # #save model
            # model_name = datetime.now().strftime("%d-%m-%Y_%H_%M_%S") + "_model_bs" + str(BATCH_SIZE) + ".pt"
            # torch.save(model.state_dict(), os.path.join('output',model_name))

            # test model
            log.info("Test model after training")
            acc_model, fold_data = test.evaluate_model(model, test_data_loader, device, log=log, fold=i + 1)
            fold_predictions.append(fold_data)

            log.info(
                f'Accuracy after training {acc_model}. | [{time.time() - t0}]')
            folds_acc.append(acc_model)

            if SAVE_LOSS:
                plt.plot(list(range(1, len(losses) + 1)), losses)
                plt.ylabel("loss")
                plt.xlabel("epochs")
                plt.title(f'loos evolution at fold {i + 1}')
                plot_name = datetime.now().strftime("%d-%m-%Y_%H_%M_%S") + "_loss_fold_" + str(i + 1) + ".png"
                log.info(plot_name)
                plt.savefig('plots/' + plot_name)

            if MONOFOLD:
                break

        # Confident interval computation
        mean, stdev, offset, ci = statistics.get_fold_metrics(folds_acc)
        log.info(f'******************************************')
        log.info(f'Model performance:')
        log.info(f'     Folds Acc.: {folds_acc}')
        log.info(f'     Mean: {mean}')
        log.info(f'     Stdev: {stdev}')
        log.info(f'     Offset: {offset}')
        log.info(f'     CI:(95%) : {ci}')
        print(f'******************************************')
        print(f'Model performance:')
        print(f'     Folds Acc.: {folds_acc}')
        print(f'     Mean: {mean}')
        print(f'     Stdev: {stdev}')
        print(f'     Offset: {offset}')
        print(f'     CI:(95%) : {ci}')

        fold_predictions = [item for sublist in fold_predictions for item in sublist]

        fd = open(FOLD_PREDICTIONS, 'wb')
        pickle.dump(fold_predictions, fd)
        fd.close()

    else:

        f = open(FOLD_PREDICTIONS, 'rb')
        fold_predictions = pickle.load(f)
        f.close()

        pos=0
        for elem in fold_predictions:
            name = elem[0]
            if 'Dc2' in name:
                break
            pos+=1
        value = fold_predictions[pos]
        fold_predictions[pos]= ('p164ODc',value[1],value[2],value[3],value[4],value[5],value[6],value[7])

        fold_predictions_right={}
        fold_predictions_left={}
        for pred in fold_predictions:
            img = pred[0][0:len(pred[0])-3]
            if 'ODc' in pred[0]:
                fold_predictions_right[img] = (pred[1], pred[2],pred[3], pred[4], pred[5], pred[6], pred[7])
            if 'OIc' in pred[0]:
                fold_predictions_left[img] = (pred[1], pred[2],pred[3], pred[4], pred[5],pred[6],pred[7])

        for pred in fold_predictions:
            img = pred[0][0:len(pred[0])-3]
            assert fold_predictions_right[img][5] == fold_predictions_left[img][5]
            assert fold_predictions_right[img][6] == fold_predictions_left[img][6]

        import openpyxl
        wb = openpyxl.reader.excel.load_workbook(BASE_CLINICAL)

        todelete={}
        for sheet in wb.worksheets:
            sheet['E1'] = 'P_left_eye'
            sheet['F1'] = 'P_right_eye'
            sheet['G1'] = 'P_SF1_left_eye'
            sheet['H1'] = 'P_SF2_left_eye'
            sheet['I1'] = 'P_SF1_right_eye'
            sheet['J1'] = 'P_SF2_right_eye'
            sheet['K1'] = 'Prediction_left'
            sheet['L1'] = 'Label_left'
            sheet['M1'] = 'Prediction_right'
            sheet['N1'] = 'Label_right'
            sheet['O1'] = 'fold'
            sheet['P1'] = 'class_name'
            todelete[sheet.title] = []
            for f in sheet.iter_rows(sheet.min_row, sheet.max_row):
                if f[0].value in fold_predictions_left.keys():
                    sheet[f'E{f[0].row}'] = fold_predictions_left[f[0].value][0]
                    sheet[f'G{f[0].row}'] = fold_predictions_left[f[0].value][1]
                    sheet[f'H{f[0].row}'] = fold_predictions_left[f[0].value][2]
                    sheet[f'K{f[0].row}'] = fold_predictions_left[f[0].value][3]
                    sheet[f'L{f[0].row}'] = fold_predictions_left[f[0].value][4]
                    sheet[f'O{f[0].row}'] = fold_predictions_left[f[0].value][5]
                    sheet[f'P{f[0].row}'] = fold_predictions_left[f[0].value][6]
                else:
                    if f[0].row > 1:
                        todelete[sheet.title].append(f[0].row)


                if f[0].value in fold_predictions_right.keys():
                    sheet[f'F{f[0].row}'] = fold_predictions_right[f[0].value][0]
                    sheet[f'I{f[0].row}'] = fold_predictions_right[f[0].value][1]
                    sheet[f'J{f[0].row}'] = fold_predictions_right[f[0].value][2]
                    sheet[f'M{f[0].row}'] = fold_predictions_right[f[0].value][3]
                    sheet[f'N{f[0].row}'] = fold_predictions_right[f[0].value][4]
                else:
                    if f[0].row > 1:
                        todelete[sheet.title].append(f[0].row)
        for sheet, rows in todelete.items():
            for r in reversed(rows):
                wb[sheet].delete_rows(r)

        wb.save(CLINICAL_PREDICTION)

        # TMP Generate file by fold
        wb_source = openpyxl.reader.excel.load_workbook(CLINICAL_PREDICTION)
        for i in range(5):
            wb_target = openpyxl.reader.excel.load_workbook(CLINICAL_PREDICTION)
            fold = i+1
            filename = f'data/clinical/clinical_data_prediction_fold_{fold}.xlsx'

            todelete={}
            for sheet in wb_target.worksheets:
                todelete[sheet.title] = []
                for f in sheet.iter_rows(sheet.min_row, sheet.max_row):
                    if f[14].value != fold and f[14].row > 1:
                        todelete[sheet.title].append(f[0].row)

            for sheet, rows in todelete.items():
                for r in reversed(rows):
                    wb_target[sheet].delete_rows(r)


            wb_target.save(filename)




        '''
        ##################################################################################################
        Remove RD from continuous clinical data
        '''
        total = ['P_left_eye', 'P_right_eye', 'P_SF1_left_eye',
                 'P_SF2_left_eye',
                 'P_SF1_right_eye',
                 'P_SF2_right_eye',
                 'Prediction_left',
                 'Label_left',
                 'Prediction_right',
                 'Label_right',
                 'fold',
                 'class_name']

        # log.info('')
        # log.info('Running classifiers with discrete prediction data')
        #
        # for i in range(5):
        #     fold = i+1
        #     filename = f'data/clinical/clinical_data_prediction_fold_{fold}.xlsx'
        #     log.info(f'Running classifiers for fold {fold}')
        #     results, avg = run_classifiers(filename, ['edat', 'sexe', 'Rd.si/no', 'P_SF1_left_eye',
        #              'P_SF2_left_eye',
        #              'P_SF1_right_eye',
        #              'P_SF2_right_eye',])
        #     log.info(results)
        #     log.info('--------------')
        #     log.info(f'Mean classifiers: {avg}')
        #     log.info('--------------')
        #     log.info('')


        log.info('')
        log.info('Running classifiers with discrete prediction data')


        results, avg = run_classifiers(CLINICAL_PREDICTION, ['edat', 'sexe', 'Rd.si/no'])
        log.info(results)
        log.info('--------------')
        log.info(f'Mean classifiers: {avg}')
        log.info('--------------')
        log.info('')


